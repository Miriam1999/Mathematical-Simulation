import numpy as np
from math import exp
from csv import reader

import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook



def xlsx2array(filename, sheet, initCel, endCel):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb[sheet]
    
    # Read the cell values into a list of lists
    data_rows = []
    for row in ws[initCel : endCel]:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)
    return(data_rows)



#Input array
X = np.array(xlsx2array('PS.xlsx', 'PS', 'A4','E14'))
#print(X)
'''X=np.array([[0,0,0,0,1],
            [0,0,0,1,1],
			[0,0,1,0,1],
			[0,0,1,1,1],
			[0,1,0,1,1],
			[0,1,1,0,1],
			[0,1,1,1,1],
            [1,0,0,1,1],
            [1,0,1,0,1],
            [1,1,0,1,1],
            [1,1,1,0,1]])
'''

#Output
y = np.array(xlsx2array('PS.xlsx', 'PS', 'F4','F14'))
#print(y)
'''y=np.array([[1],
            [1],
			[1],
            [0],
            [1],
            [0],
			[1],
			[0],
			[0],
			[0],
            [1]])
'''
#%%





#Sigmoid Function
def sigmoid (x):
    return 1/(1 + np.exp(-x))

#Derivative of Sigmoid Function
def derivatives_sigmoid(x):
    return x * (1 - x)

#Variable initialization
epoch               = 5000        #Setting training iterations
lr                  = 0.1         #Setting learning rate
inputlayer_neurons  = X.shape[1] #number of features in data set
hiddenlayer_neurons = 5          #number of hidden layers neurons
output_neurons      = y.shape[1] #number of neurons at output layer

#weight and bias initialization
wh   = np.random.uniform(size=(inputlayer_neurons,hiddenlayer_neurons))
bh   = np.random.uniform(size=(1,hiddenlayer_neurons))
wout = np.random.uniform(size=(hiddenlayer_neurons,output_neurons))
bout = np.random.uniform(size=(1,output_neurons))
#%%
for i in range(epoch):
    
    #Forward Propogation
    hidden_layer_input1     = np.dot(X,wh)
    hidden_layer_input      = hidden_layer_input1 + bh
    hiddenlayer_activations = sigmoid(hidden_layer_input)
    output_layer_input1     = np.dot(hiddenlayer_activations,wout)
    output_layer_input      = output_layer_input1+ bout
    output                  = sigmoid(output_layer_input)
    
    #Backpropagation
    E                     = y-output
    slope_output_layer    = derivatives_sigmoid(output)
    slope_hidden_layer    = derivatives_sigmoid(hiddenlayer_activations)
    d_output              = E * slope_output_layer
    Error_at_hidden_layer = d_output.dot(wout.T)
    d_hiddenlayer         = Error_at_hidden_layer * slope_hidden_layer
    
    wout += hiddenlayer_activations.T.dot(d_output) *lr
    bout += np.sum(d_output, axis=0,keepdims=True) *lr
    wh   += X.T.dot(d_hiddenlayer) *lr
    bh   += np.sum(d_hiddenlayer, axis=0,keepdims=True) *lr

print ("\n","-------------------- Weights Hidden Layer --------------------")
print (wh)
print ("\n","-------------------- Weights Output Layer --------------------")
print (wout)
print ("\n","--------------------         Output       --------------------")
print (output)











